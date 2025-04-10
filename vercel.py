from flask import Flask, render_template, request, jsonify, send_file
import os
import sys
import json
import tempfile
import traceback
import time
import requests
from werkzeug.utils import secure_filename

# 设置日志
import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    stream=sys.stdout
)
logger = logging.getLogger('vercel')

# 记录环境信息
logger.info(f"Starting Vercel-compatible application")
logger.info(f"Python version: {sys.version}")
logger.info(f"Working directory: {os.getcwd()}")
logger.info(f"Files in current directory: {os.listdir('.')}")

# 创建Flask应用
app = Flask(__name__)
app.config['SECRET_KEY'] = 'weread-exporter-secret-key!'

# 在Vercel上使用/tmp目录，因为它是唯一可写的
UPLOAD_FOLDER = '/tmp/uploads' if os.environ.get('VERCEL') == '1' else 'uploads'
OUTPUT_DIR = '/tmp/outputs' if os.environ.get('VERCEL') == '1' else 'outputs'

try:
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
        logger.info(f"Created upload folder: {UPLOAD_FOLDER}")

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logger.info(f"Created output folder: {OUTPUT_DIR}")
except Exception as e:
    logger.error(f"Error creating directories: {str(e)}")

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB限制

# 检查openpyxl是否可用（不使用pandas）
try:
    import openpyxl
    from openpyxl import Workbook
    has_excel_support = True
    logger.info("Successfully imported openpyxl")
except ImportError:
    has_excel_support = False
    logger.warning("openpyxl not available, Excel export will be disabled")

# 从notebook_v1.py中提取的必要函数，移除pandas依赖
def parse_cookie_string(cookie_string):
    """解析cookie字符串为Cookie对象"""
    from http.cookies import SimpleCookie
    cookie = SimpleCookie()
    cookie.load(cookie_string)
    cookies_dict = {}
    
    for key, morsel in cookie.items():
        cookies_dict[key] = morsel.value
    
    return cookies_dict

def get_notebooklist(session):
    """获取笔记本列表"""
    url = "https://i.weread.qq.com/user/notebooks"
    response = session.get(url)
    
    if response.status_code == 200:
        return response.json().get('notebooks', [])
    return []

def get_bookinfo(session, bookId):
    """获取书籍信息"""
    url = f"https://i.weread.qq.com/book/info?bookId={bookId}"
    response = session.get(url)
    
    if response.status_code == 200:
        data = response.json()
        isbn = data.get('isbn', '')
        rating = data.get('newRating', 0)
        return isbn, rating, data
    return "", 0, {}

def get_chapter_info(session, bookId):
    """获取章节信息"""
    url = f"https://i.weread.qq.com/book/chapterInfos?bookId={bookId}"
    response = session.get(url)
    
    if response.status_code == 200:
        data = response.json()
        chapters = data.get('data', [])
        chapter_info = {}
        
        for chapter in chapters:
            chapter_info[chapter.get('chapterUid')] = chapter
        
        return chapter_info
    return {}

def get_bookmark_list(session, bookId):
    """获取划线列表"""
    url = f"https://i.weread.qq.com/book/bookmarklist?bookId={bookId}"
    response = session.get(url)
    
    if response.status_code == 200:
        data = response.json()
        return data.get('updated', [])
    return []

def get_review_list(session, bookId):
    """获取笔记列表"""
    url = f"https://i.weread.qq.com/review/list?bookId={bookId}"
    response = session.get(url)
    
    if response.status_code == 200:
        data = response.json()
        summary = data.get('summary', {})
        reviews = data.get('reviews', [])
        return summary, reviews
    return {}, []

def export_to_json(books_data, output_file):
    """导出为JSON格式"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(books_data, f, ensure_ascii=False, indent=4)
    return True

def export_to_excel(books_data, output_file):
    """导出为Excel格式，仅使用openpyxl，不依赖pandas"""
    if not has_excel_support:
        logger.warning("openpyxl not available, cannot export to Excel")
        return False
    
    try:
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "微信读书笔记"
        
        # 添加表头
        headers = ['书名', '作者', 'ISBN', '评分', '类型', '章节', '创建时间', '内容']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
        
        # 添加所有笔记数据
        row_num = 2
        for book_data in books_data:
            book_info = book_data.get('book_info', {})
            book_title = book_info.get('title', '未知书名')
            book_author = book_info.get('author', '未知作者')
            isbn = book_data.get('isbn', '')
            rating = book_data.get('rating', 0)
            
            notes = book_data.get('notes', [])
            if not notes:
                continue
                
            # 处理每条笔记
            for note in notes:
                note_type = '划线' if note.get('type') == 1 else '笔记'
                chapter = note.get('chapter_title', '')
                created_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(note.get('createTime', 0)))
                content = note.get('markText', '') or note.get('content', '')
                
                # 写入每一列
                col_num = 1
                for value in [book_title, book_author, isbn, rating, note_type, chapter, created_time, content]:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    col_num += 1
                
                row_num += 1
        
        # 调整列宽以适应内容
        for col_num, _ in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            if col_num < len(headers):
                ws.column_dimensions[col_letter].width = 20
            else:
                # 内容列宽度设置得更大一些
                ws.column_dimensions[col_letter].width = 50
        
        # 保存工作簿
        wb.save(output_file)
        logger.info(f"Excel exported successfully: {output_file}")
        return True
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        logger.error(traceback.format_exc())
        return False

@app.route('/')
def index():
    logger.info("Serving index page")
    return render_template('index.html')

@app.route('/status', methods=['GET'])
def status():
    """状态检查接口"""
    logger.info("Status check called")
    return jsonify({
        'status': 'ok',
        'version': '1.0.0 (Vercel)',
        'environment': 'vercel' if os.environ.get('VERCEL') == '1' else 'local',
        'python_version': sys.version,
        'directories': {
            'uploads': os.path.exists(UPLOAD_FOLDER),
            'outputs': os.path.exists(OUTPUT_DIR)
        },
        'excel_support': has_excel_support
    })

@app.route('/extract', methods=['POST'])
def extract():
    logger.info("Extract endpoint called")
    try:
        # 获取cookie
        cookie = request.form.get('cookie', '')
        
        if not cookie:
            logger.warning("No cookie provided")
            return jsonify({'status': 'error', 'message': '请输入有效的Cookie'}), 400
            
        # 创建临时目录用于存储导出文件
        try:
            temp_dir = tempfile.mkdtemp(dir=OUTPUT_DIR)
            logger.info(f"Created temp directory: {temp_dir}")
        except Exception as e:
            logger.error(f"Error creating temp directory: {str(e)}")
            temp_dir = OUTPUT_DIR
        
        # 获取用户的User-Agent
        user_agent = request.headers.get('User-Agent', '')
        if not user_agent:
            user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        
        # 创建会话
        session = requests.Session()
        session.cookies = parse_cookie_string(cookie)
        session.headers.update({'User-Agent': user_agent})
        
        # 访问主页获取必要的cookie
        try:
            weread_url = "https://weread.qq.com/"
            logger.info(f"Accessing weread URL: {weread_url}")
            response = session.get(weread_url, timeout=5)
            logger.info(f"Weread response status: {response.status_code}")
        except Exception as e:
            logger.error(f"Failed to access weread: {str(e)}")
            return jsonify({'status': 'error', 'message': f'访问微信读书主页失败: {str(e)}'}), 500
        
        # 获取笔记本列表
        try:
            logger.info("Fetching notebook list")
            books = get_notebooklist(session)
            if not books:
                logger.warning("No books found")
                return jsonify({'status': 'error', 'message': '获取书籍列表失败，请检查Cookie是否有效'}), 400
            
            logger.info(f"Found {len(books)} books")
        except Exception as e:
            logger.error(f"Error fetching notebook list: {str(e)}")
            return jsonify({'status': 'error', 'message': f'获取书籍列表失败: {str(e)}'}), 500
        
        # 限制处理的书籍数量以避免超时
        max_books = 15  # 限制每次处理的最大书籍数量
        if len(books) > max_books and os.environ.get('VERCEL') == '1':
            logger.warning(f"Too many books ({len(books)}), limiting to {max_books} to avoid timeout")
            books = books[:max_books]
            
        # 发送总书籍数量
        total_books = len(books)
        all_books_data = []
        
        for index, book_item in enumerate(books):
            try:
                # 检查是否运行在Vercel环境，如果是，则检查是否快要超时
                if os.environ.get('VERCEL') == '1' and time.time() - request.environ.get('FLASK_REQUEST_START_TIME', time.time()) > 8:
                    logger.warning("Request is about to timeout, stopping processing")
                    break
                    
                book = book_item.get('book')
                bookId = book.get('bookId')
                title = book.get('title', '未知书名')
                
                # 更新进度
                current_book = index + 1
                percent = int((current_book / total_books) * 100)
                logger.info(f"Processing book {current_book}/{total_books}: {title}")
                
                # 获取书籍详细信息
                isbn, rating, book_info = get_bookinfo(session, bookId)
                chapter_info = get_chapter_info(session, bookId)
                
                # 划线列表清洗
                bookmark_list = get_bookmark_list(session, bookId)
                if bookmark_list:
                    bookmark_list_huaxian = [item for item in bookmark_list if item.get('type') == 1]
                else:
                    bookmark_list_huaxian = []
                    
                # 书评和笔记
                summary, reviews = get_review_list(session, bookId)
    
                # 合并划线和笔记
                all_notes = []
                if bookmark_list_huaxian:
                    all_notes.extend(bookmark_list_huaxian)
                    logger.info(f"Book '{title}' has {len(bookmark_list_huaxian)} highlights")
                if reviews:
                    all_notes.extend(reviews)
                    logger.info(f"Book '{title}' has {len(reviews)} notes")
                
                # 排序
                if all_notes:
                    all_notes = sorted(
                        all_notes,
                        key=lambda x: (
                            x.get("chapterUid", 1),
                            (
                                0
                                if (
                                    x.get("range", "") == ""
                                    or x.get("range", "").split("-")[0] == ""
                                )
                                else int(x.get("range", "0-0").split("-")[0])
                            ),
                        ),
                    )
                
                # 添加章节信息
                for note in all_notes:
                    chapterUid = note.get("chapterUid", 1)
                    if chapter_info and chapterUid in chapter_info:
                        note["chapter_title"] = chapter_info[chapterUid].get("title", "")
                    else:
                        note["chapter_title"] = ""
                
                # 添加书籍数据
                book_data = {
                    "book_info": book,
                    "isbn": isbn,
                    "rating": rating,
                    "notes": all_notes,
                    "summary": summary
                }
                
                all_books_data.append(book_data)
                
                # 每处理一本书睡眠很短时间，避免请求过快但不会明显延长总处理时间
                time.sleep(0.2)
            except Exception as e:
                logger.error(f"Error processing book '{title}': {str(e)}")
                logger.error(traceback.format_exc())
                # 继续处理下一本书
                continue
        
        # 导出数据
        timestamp = int(time.time())
        json_file = os.path.join(temp_dir, f'weread_notes_{timestamp}.json')
        
        logger.info(f"Exporting data to JSON: {json_file}")
        export_to_json(all_books_data, json_file)
        
        # 尝试导出Excel
        response_data = {
            'status': 'success', 
            'message': '数据导出成功',
            'files': {
                'json': f'/download?file=weread_notes_{timestamp}.json&dir={os.path.basename(temp_dir)}'
            }
        }
        
        if has_excel_support:
            try:
                excel_file = os.path.join(temp_dir, f'weread_notes_{timestamp}.xlsx')
                logger.info(f"Exporting data to Excel: {excel_file}")
                if export_to_excel(all_books_data, excel_file):
                    response_data['files']['excel'] = f'/download?file=weread_notes_{timestamp}.xlsx&dir={os.path.basename(temp_dir)}'
                    logger.info("Excel export successful")
                else:
                    logger.warning("Excel export failed")
            except Exception as e:
                logger.error(f"Error during Excel export: {str(e)}")
        else:
            response_data['note'] = '当前环境不支持Excel导出，仅提供JSON格式'
            
        # 添加处理信息
        if len(books) > len(all_books_data):
            if os.environ.get('VERCEL') == '1':
                response_data['warning'] = f'由于Vercel环境限制，仅处理了{len(all_books_data)}本书中的{len(all_books_data)}本。建议在本地环境运行以获取所有数据。'
            else:
                response_data['warning'] = f'仅处理了{len(books)}本书中的{len(all_books_data)}本，有些书籍处理失败。'
        
        logger.info(f"Processing completed successfully - processed {len(all_books_data)} of {len(books)} books")
        
        return jsonify(response_data)
        
    except Exception as e:
        error_msg = traceback.format_exc()
        logger.error(f"Extraction failed: {str(e)}")
        logger.error(error_msg)
        return jsonify({'status': 'error', 'message': f'处理过程中出错: {str(e)}', 'details': error_msg}), 500

@app.route('/download')
def download():
    logger.info("Download endpoint called")
    filename = request.args.get('file')
    dir_name = request.args.get('dir')
    
    if not filename or not dir_name:
        logger.warning("Invalid download parameters")
        return jsonify({'status': 'error', 'message': '无效的下载参数'}), 400
    
    # 安全检查
    filename = secure_filename(filename)
    if '..' in dir_name or '/' in dir_name:
        logger.warning(f"Invalid directory parameter: {dir_name}")
        return jsonify({'status': 'error', 'message': '无效的目录参数'}), 400
    
    file_path = os.path.join(OUTPUT_DIR, dir_name, filename)
    logger.info(f"Download file path: {file_path}")
    
    if not os.path.exists(file_path):
        logger.warning(f"File not found: {file_path}")
        return jsonify({'status': 'error', 'message': '文件不存在'}), 404
    
    return send_file(file_path, as_attachment=True)

# 记录请求开始时间以便计算超时
@app.before_request
def before_request():
    request.environ['FLASK_REQUEST_START_TIME'] = time.time()

# 导出app对象供Vercel使用
application = app

if __name__ == '__main__':
    app.run(debug=True) 